import openpyxl
import os
from PyPDF2 import PdfWriter
import time
from openpyxl.styles import Color, PatternFill, Font, Border

def combinePDF():
    for maindoc in my_dict.keys():
        try:
            if os.path.isfile(os.path.join(target, maindoc+".pdf")):
                continue
            else:
                merger = PdfWriter()
                print("working on " +maindoc ,end="\r")
                for child in my_dict[maindoc]:                     #this part need to be checked since the dictionary is modified
                    merger.append(os.path.join(source,child+".pdf"))
                merger.write(os.path.join(target, maindoc+".pdf"))
                merger.close()
                print(maindoc, "is done                          ")
        except Exception as e:
            second = time.time()
            f=open(str(second)+"log.txt",'w')
            f.write("\n")
            f.write(str(e))
            f.close()

def get_sus_filename(file):
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    count = 0
    flag = 0
    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        if str(cell_obj.value).startswith("None"):
            continue
        else:
            if sheet_obj.cell(row=i, column=5).value >= 10 or "(1" in (str(cell_obj.value)):
                if flag == 0:
                    print(file)
                    flag = 1
                else:
                    print("              ", str(cell_obj.value),"              files: ", sheet_obj.cell(row=i, column=5).value)

def read_excel(file):
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    count = 0
    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=2)
        if str(cell_obj.value).startswith("None"):
            continue
        else:
            if os.path.isfile(str(cell_obj.value)[1:-1]):
                continue
            else:
                print(str(sheet_obj.cell(row=i, column=2).value)[1:-1])
                for child in range(0, sheet_obj.cell(row=i, column=5).value):
                    print("                                  ",str(sheet_obj.cell(row=i+child, column=4).value)[1:-1])
                print()
                print()
                count += 1
    print(count)


def get_key(file):
    key = ""
    for i in reversed(file):
        if i == '-':
           key = i
           break
        elif i == '.':
            key = i
            break
        elif i == ' ':
            key = i
            break
        elif i == '/':
            key = i
            break
        elif i == '#':
            key = i
            break
        elif i == '_':
            key = i
            break
    if key:
        x = file.split(key)
        after = key.join(x[:-1])
        return after
    return file

def addToDict(source):
    for root, dirs, files in os.walk(source):
        count = 0
        for index in range(len(files) - 1):
            file = files[index][:-4]  # remove ".pdf"
            shorten = get_key(file)
            if file not in files[index + 1][:-4]:
                if shorten not in my_dict:
                    my_dict[shorten] = [file]
                else:
                    my_dict[shorten].append(file)
            else:
                my_dict[file] = [file]





my_dict = {}
source = r"\\FTPCNTSK\Images\Fortis Export"
target = r"\\FTPCNTSK\Images\FortisMerged"



for root,dir,files in os.walk("./"):
    for file in files:
        if file.endswith(".xlsx"):
            get_sus_filename(file)







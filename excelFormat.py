import openpyxl
import os
from openpyxl.styles import PatternFill
def makeTable(input, my_dict ,source,target):
    dir_sheet = openpyxl.Workbook()
    dir_data = dir_sheet.active
    r = 2
    dir_data.cell(row=1, column=1, value="merged files(parent)")
    dir_data.cell(row=1, column=2, value="link")
    dir_data.cell(row=1, column=3, value="files(child)")
    dir_data.cell(row=1, column=4, value="link(child)")
    dir_data.cell(row=1, column=5, value="# of Children")
    for key in my_dict.keys():
        path = os.path.join(target, key)
        dir_data.cell(row=r, column=1, value=key)
        dir_data.cell(row=r, column=1).hyperlink = path+".pdf"
        dir_data.cell(row=r, column=2, value = "#" + path + ".pdf#")
        dir_data.cell(row=r, column=5, value=len(my_dict[key]))
        for x in my_dict[key]:
            path = os.path.join(source, x)
            dir_data.cell(row=r, column=3, value=x)
            dir_data.cell(row=r, column=3).hyperlink = path+".pdf"
            dir_data.cell(row=r, column=4, value = "#"+path + ".pdf#")
            r+=1
    dir_data.column_dimensions['A'].width = 30              #for width
    dir_data.column_dimensions['B'].width = 1
    dir_data.column_dimensions['C'].width = 30
    dir_data.column_dimensions['D'].width = 1
    dir_data.column_dimensions['E'].width = 10

    dir_data.freeze_panes = 'A2'

    dir_data.cell(row = 1, column = 1).fill = PatternFill(start_color='00FFFF',end_color='00FFFF',fill_type='solid')    #for color on top
    dir_data.cell(row = 1, column = 2).fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    dir_data.cell(row = 1, column = 3).fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    dir_data.cell(row = 1, column = 4).fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    dir_data.cell(row = 1, column = 5).fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    dir_data.cell(row = 1, column = 6).fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')

    dir_sheet.save(input+"_Merged_Table.xlsx")
    dir_sheet.close()